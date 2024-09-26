from openpyxl import load_workbook

import consts
import go_const_template


class Parse:
    class Data:
        def __init__(self):
            self.desc_name = ""
            self.field_name = ""
            self.field_type = ""
            self.field_value = ""

    def __init__(self):
        self.data_list = []

    def proc_main(self):
        self.parse()

        self.gen_server_const_code()

    def gen_server_const_code(self):
        content = ""
        for data in self.data_list:
            desc_name = data.desc_name
            field_value = data.field_value
            field_name = data.field_name[0].upper() + data.field_name[1:]
            field_type = data.field_type

            code = f"{field_name} = {field_value} // {desc_name}"
            if field_type != "str" or field_type != "string":
                code = f"{field_name} = {field_type}({field_value}) // {desc_name}"

            content += code

        with open(consts.server_output_dist, "w", encoding="utf-8") as f:
            file_content = go_const_template.template
            file_content = file_content.replace(go_const_template.content_replace, content)

            f.write(file_content)

    def parse(self):
        workbook = load_workbook(consts.input_data)
        sheet = workbook.active

        # 遍历每一行
        for row in sheet.iter_rows():
            # print(f"type {type(row)}")
            data = self.Data()
            data.desc_name = row[0].value
            data.field_value = row[1].value
            data.field_name = row[2].value
            data.field_type = row[3].value

            if data.field_type == "str" or data.field_type == "string":
                data.field_value = str(data.field_value)
            self.data_list.append(data)
            # print(data)


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    print("start run gen const code")
    parse = Parse()
    parse.proc_main()
    print("end run gen const code")
